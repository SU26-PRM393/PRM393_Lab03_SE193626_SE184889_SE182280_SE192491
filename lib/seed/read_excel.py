import openpyxl
import json

file_path = "d:/Semesters/Summer2026/PRM393/Project/VietNam-Map-Flutter/Truong_THPT_2026_import_ready.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

data = []
headers = [cell.value for cell in ws[1]]

for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(row): continue # skip empty rows
    data.append(dict(zip(headers, row)))

with open('schools.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"Exported {len(data)} schools to schools.json")
print("First 2 schools:")
print(json.dumps(data[:2], ensure_ascii=False, indent=2))
